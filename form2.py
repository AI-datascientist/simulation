# -*- coding: utf-8 -*-
# Psychiatric Interview Simulation — Streamlit (Voice/Text, Talking Avatars, Sidebar Info)
# Optimized: cached background & avatars, cached heavy resources, persistent mic (no reruns),
# limited history rendering, shortened sleeps. Everything in English.

import os, re, uuid, json, datetime, time, base64
from typing import List, Dict, Any, Tuple, Optional
from io import BytesIO
from difflib import SequenceMatcher

import streamlit as st
import streamlit.components.v1 as components
from PIL import Image

# =============================
# OPTIONAL DEPS (gracefully degrade)
# =============================
try:
    import google.generativeai as genai
    _GENAI_IMPORT_OK = True
except Exception:
    _GENAI_IMPORT_OK = False

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    _XLSX_OK = True
except Exception:
    _XLSX_OK = False

try:
    from audio_recorder_streamlit import audio_recorder
    _RECORDER_OK = True
except Exception:
    _RECORDER_OK = False

try:
    import SpeechRecognition as sr  # package name is SpeechRecognition; module is speech_recognition
except Exception:
    try:
        import speech_recognition as sr
        _SR_OK = True
    except Exception:
        _SR_OK = False
else:
    _SR_OK = True

# =============================
# CONFIG
# =============================
APP_TITLE = "Psychiatric Interview Simulation"
LOG_DIR = "logs"
USER_DIR = "users"
os.makedirs(LOG_DIR, exist_ok=True)
os.makedirs(USER_DIR, exist_ok=True)

EXCEL_FILE = os.path.join(USER_DIR, "registered_users.xlsx")
EXCEL_PASSWORD = "admin123"
DOWNLOAD_PASSWORD = "download456"

MAX_MSG = 30  # render only last N messages to keep UI fast

# Prefer env or secrets; allow UI entry
_GOOGLE_API_KEY_ENV = os.getenv("GOOGLE_API_KEY", "").strip()
try:
    _GOOGLE_API_KEY_SEC = st.secrets.get("GOOGLE_API_KEY", "").strip()
except Exception:
    _GOOGLE_API_KEY_SEC = ""

# =============================
# PAGE SETUP
# =============================
st.set_page_config(
    page_title=APP_TITLE,
    page_icon="👥",
    layout="wide",
    initial_sidebar_state="expanded"
)

# =============================
# CACHED HELPERS (images, css)
# =============================
@st.cache_data(show_spinner=False)
def load_image_b64(path: str) -> str:
    if not os.path.exists(path):
        return ""
    with open(path, "rb") as f:
        return base64.b64encode(f.read()).decode()

@st.cache_data(show_spinner=False)
def background_css_from_b64(b64: str) -> str:
    return f"""
    <style>
    .stApp {{
        background-image: url("data:image/jpg;base64,{b64}");
        background-size: cover;
        background-position: center;
        background-repeat: no-repeat;
        background-attachment: fixed;
    }}
    .chat-bubble {{
        background: rgba(255,255,255,0.92);
        border-radius: 14px;
        padding: 12px 14px;
        border: 1px solid rgba(0,0,0,0.06);
    }}
    </style>
    """

def set_background_from_file(path: str):
    b64 = load_image_b64(path)
    if b64:
        st.markdown(background_css_from_b64(b64), unsafe_allow_html=True)

@st.cache_data(show_spinner=False)
def load_avatar_b64(path: str) -> str:
    return load_image_b64(path)

set_background_from_file("bck.webp" if os.path.exists("bck.webp") else "bck.jpg")

# =============================
# PERSONAS
# =============================
MDD_PERSONA = {
    "name": "Aliye Seker",
    "age": 40,
    "gender": "Female",
    "dx": "Major Depressive Disorder",
    "current_meds": "Fluoxetine 20–40 mg (day 7 of uptitration)",
    "photo": "aliye.webp" if os.path.exists("aliye.webp") else "aliye.jpg",
    "speech_style": "Brief answers, slow, hopeless tone."
}
SCZ_PERSONA = {
    "name": "Feride Deniz",
    "age": 25,
    "gender": "Female",
    "dx": "Schizophrenia, Paranoid Type",
    "current_meds": "LAI Risperidone; previously Haloperidol",
    "photo": "feride.webp" if os.path.exists("feride.webp") else "feride.jpg",
    "speech_style": "Guarded, may be tangential; paranoid themes."
}

# =============================
# SIMPLE STANDARDIZED QA DB (demo)
# =============================
ALI_QA_DATABASE = [
    {"q": "How are you feeling today", "a": "I still feel like I'm in a dark hole. Nothing seems to help.", "part": "Part 1"},
    {"q": "Do you have thoughts of harming yourself", "a": "I've had those thoughts… but I don't want to go into details.", "part": "Part 1"},
    {"q": "How is your sleep", "a": "I wake up around 3 AM and can't go back to sleep.", "part": "Part 1"},
]
FERDI_QA_DATABASE = [
    {"q": "Why are you here", "a": "My mother brought me. She thinks I need help.", "part": "Part 1"},
    {"q": "Do you hear voices", "a": "Sometimes… but they're quieter when I stay calm.", "part": "Part 2"},
]

# =============================
# STATE / UTIL
# =============================
def ensure_state_defaults():
    ss = st.session_state
    ss.setdefault("page", "registration")
    ss.setdefault("user", None)
    ss.setdefault("selected_persona", None)
    ss.setdefault("selected_part", None)
    ss.setdefault("session_id", None)
    ss.setdefault("conversation", [])
    ss.setdefault("awaiting_permission", False)
    ss.setdefault("avatar_placeholder", None)
    ss.setdefault("enable_tts", True)
    ss.setdefault("voice_output_target", "Browser (SpeechSynthesis)")
    ss.setdefault("GOOGLE_API_KEY_UI", "")
    ss.setdefault("pending_voice_input", "")
    ss.setdefault("recorder_key", f"rec_{uuid.uuid4().hex[:8]}")  # stable mic key

ensure_state_defaults()

def timestamp():
    return datetime.datetime.now().strftime("%H:%M:%S")

def log_line(role, text, source=""):
    st.session_state.conversation.append((role, text, timestamp(), source))

def similarity(a: Any, b: Any) -> float:
    a = str(a or "")
    b = str(b or "")
    a = re.sub(r"[^\w\s]", "", a.lower().strip())
    b = re.sub(r"[^\w\s]", "", b.lower().strip())
    try:
        return SequenceMatcher(None, a, b).ratio()
    except Exception:
        return 0.0

def qa_lookup(user_q: Any, persona_name: str, part: str) -> Optional[str]:
    user_q = str(user_q or "")
    db = ALI_QA_DATABASE if persona_name == "Aliye Seker" else FERDI_QA_DATABASE
    best, best_s = None, 0.0
    for row in db:
        if row.get("part") == part:
            s = similarity(user_q, row.get("q", ""))
            if s > best_s:
                best_s, best = s, row
    if best and best_s >= 0.70:
        return best.get("a", "")
    return None

# =============================
# GEMINI (cached heavy resource)
# =============================
def get_google_api_key() -> str:
    if _GOOGLE_API_KEY_ENV:
        return _GOOGLE_API_KEY_ENV
    if _GOOGLE_API_KEY_SEC:
        return _GOOGLE_API_KEY_SEC
    return st.session_state.get("GOOGLE_API_KEY_UI", "").strip()

@st.cache_resource(show_spinner=False)
def get_gemini_model(system_prompt: str):
    api_key = get_google_api_key()
    if not (_GENAI_IMPORT_OK and api_key):
        return None
    genai.configure(api_key=api_key)
    return genai.GenerativeModel(model_name="gemini-2.0-flash", system_instruction=system_prompt)

def build_system_prompt(persona: Dict, part: str) -> str:
    if persona["name"] == "Aliye Seker":
        stage = "Day 2 acute admission" if part == "Part 1" else "Day 7 reassessment"
        return (
            f"You are {persona['name']}, {persona['age']}, {persona['gender']} with severe depression. "
            f"Stage: {stage}. Speak briefly, slow, hopeless. Avoid specific self-harm methods. "
            f"Stay in character and answer in 1–3 short sentences."
        )
    else:
        stage = "Day 3 acute psychosis" if part == "Part 1" else "Day 14 stabilizing"
        return (
            f"You are {persona['name']}, {persona['age']}, {persona['gender']} with paranoid schizophrenia. "
            f"Stage: {stage}. Be guarded; mild tangentiality allowed; no medical advice. "
            f"Stay in character and answer in 1–3 short sentences."
        )

def fallback_reply(persona: Dict, part: str, user_text: str) -> str:
    name = persona["name"]
    if name == "Aliye Seker":
        bank_part1 = [
            "It’s hard to feel anything. Most days are just heavy.",
            "I don’t have much energy… getting out of bed is a struggle.",
            "I sleep poorly and wake up early, feeling worse."
        ]
        bank_part2 = [
            "A little different… but the sadness is still there.",
            "I can do small things again, but it feels empty.",
            "I’m trying, but I still feel numb."
        ]
        bank = bank_part1 if part == "Part 1" else bank_part2
    else:
        bank_part1 = [
            "I don’t really trust this place… the food doesn’t feel safe.",
            "People look at me like they know my thoughts.",
            "I can talk but I’d rather keep to myself."
        ]
        bank_part2 = [
            "It’s quieter now… I can think a bit clearer.",
            "I still feel watched sometimes, so I’m careful.",
            "I’m taking the meds. It helps me stay calm."
        ]
        bank = bank_part1 if part == "Part 1" else bank_part2
    idx = abs(hash(user_text)) % len(bank) if user_text else 0
    return bank[idx]

def llm_reply(persona: Dict, part: str, user_text: str) -> str:
    sys_prompt = build_system_prompt(persona, part)
    model = get_gemini_model(sys_prompt)
    if model:
        try:
            r = model.generate_content(user_text)
            text = (getattr(r, "text", None) or "").strip()
            text = re.sub(r"\[.*?\]", "", text)
            if text:
                return text
        except Exception:
            pass
    return fallback_reply(persona, part, user_text)

# =============================
# BROWSER TTS (SpeechSynthesis)
# =============================
def speak_browser(text: str):
    escaped = (text or "").replace("\\", "\\\\").replace("`", "\\`").replace("</", "<\\/")
    components.html(
        f"""
        <script>
        const text = `{escaped}`;
        try {{
            const u = new SpeechSynthesisUtterance(text);
            u.rate = 0.9; u.pitch = 1.0; u.lang = 'en-US';
            window.speechSynthesis.cancel(); window.speechSynthesis.speak(u);
        }} catch(e) {{}}
        </script>
        """,
        height=0
    )

# =============================
# SPEECH RECOGNITION (cached recognizer)
# =============================
@st.cache_resource(show_spinner=False)
def get_sr_recognizer():
    return sr.Recognizer() if _SR_OK else None

def transcribe_wav_bytes(wav_bytes: bytes, language: str = "en-US") -> str:
    if not (_SR_OK and wav_bytes):
        return ""
    r = get_sr_recognizer()
    try:
        with sr.AudioFile(BytesIO(wav_bytes)) as source:
            audio = r.record(source)
        text = r.recognize_google(audio, language=language)  # keyless
        return (text or "").strip()
    except Exception:
        return ""

# =============================
# OPTIONAL VISUAL MIC (no return)
# =============================
def voice_input_visual_only():
    components.html(
        """
        <div style="border:3px solid #10b981;border-radius:14px;padding:16px;background:#ecfeff">
          <b>Voice Input (visual demo)</b><br>
          This microphone is visual-only. Use the press & hold button below to record.
        </div>
        """,
        height=120
    )

# =============================
# AVATAR (talk indicator)
# =============================
def show_avatar(photo_path: str, speaking: bool, placeholder=None):
    if not os.path.exists(photo_path):
        return
    if placeholder is None:
        placeholder = st.empty()
    with placeholder.container():
        st.image(photo_path, use_container_width=True)
        if speaking:
            st.markdown(
                """
                <div style='text-align:center;padding:8px;border-radius:10px;
                    background:linear-gradient(90deg,#10b981,#059669);color:#fff;
                    font-weight:600;animation:pulse 1s infinite'>Speaking...</div>
                <style>@keyframes pulse{0%,100%{opacity:1}50%{opacity:.75}}</style>
                """,
                unsafe_allow_html=True
            )
        else:
            st.markdown(
                "<div style='text-align:center;padding:8px;border-radius:10px;background:#eef2ff;color:#334155;'>Listening</div>",
                unsafe_allow_html=True
            )

# =============================
# PERSISTENT RECORDER (stable key, NO rerun)
# =============================
def persistent_recorder(language: str = "en-US") -> Optional[str]:
    """
    Always-visible mic with a stable key. Performs STT and returns transcript.
    NO st.rerun() here; caller should also avoid reruns.
    """
    if not _RECORDER_OK:
        st.error("audio-recorder-streamlit is not available. Add it to requirements.txt or use Text mode.")
        voice_input_visual_only()
        return None

    rec_container = st.container(border=True)
    with rec_container:
        st.markdown("#### 🎙️ Press & hold to record")
        audio_bytes = audio_recorder(
            text="🎙️ Press & hold to record",
            pause_threshold=1.2,
            sample_rate=16000,
            key=st.session_state.recorder_key  # stable key
        )

    if audio_bytes:
        st.audio(audio_bytes, format="audio/wav")
        text = transcribe_wav_bytes(audio_bytes, language=language)
        if text:
            st.success(f"Transcript: {text}")
            return text.strip()

    return None

# =============================
# REGISTRATION (Excel write is required)
# =============================
def save_user_profile(username, data):
    path = os.path.join(USER_DIR, f"{username}.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def load_user_profile(username):
    path = os.path.join(USER_DIR, f"{username}.json")
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return None

def save_user_to_excel(user_data):
    if not (_XLSX_OK and isinstance(user_data, dict)):
        return
    row = pd.DataFrame([{
        "Username": user_data["username"],
        "First Name": user_data["first_name"],
        "Last Name": user_data["last_name"],
        "Nickname": user_data["nickname"],
        "Email": user_data.get("email",""),
        "Registration Date": user_data["registration_date"],
        "Photo Path": user_data["photo_path"]
    }])
    if os.path.exists(EXCEL_FILE):
        try:
            old = pd.read_excel(EXCEL_FILE)
            df = pd.concat([old, row], ignore_index=True)
        except Exception:
            df = row
    else:
        df = row
    df.to_excel(EXCEL_FILE, index=False)

    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        hdr_font = Font(bold=True, color="FFFFFF")
        for c in ws[1]:
            c.fill = hdr_fill; c.font = hdr_font
            c.alignment = Alignment(horizontal="center", vertical="center")
        for col in ws.columns:
            mx = 0; letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    mx = max(mx, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(mx+2, 50)
        wb.security.workbookPassword = EXCEL_PASSWORD
        wb.security.lockStructure = True
        wb.save(EXCEL_FILE)
    except Exception:
        pass

# =============================
# UI PIECES
# =============================
def sidebar_patient_info(persona: Dict):
    st.markdown("### Patient Information")
    st.write(f"**Name:** {persona['name']}")
    st.write(f"**Age:** {persona['age']}")
    st.write(f"**Gender:** {persona['gender']}")
    st.write(f"**Diagnosis:** {persona['dx']}")
    st.write(f"**Current Meds:** {persona['current_meds']}")
    st.markdown("---")

    st.markdown("### Patient Avatar")
    if st.session_state.avatar_placeholder is None:
        st.session_state.avatar_placeholder = st.empty()
    photo_path = persona.get("photo","")
    if os.path.exists(photo_path):
        show_avatar(photo_path, speaking=False, placeholder=st.session_state.avatar_placeholder)
    else:
        st.info(f"Avatar not found: {photo_path}")

    st.markdown("---")
    st.markdown("### Settings")
    if not get_google_api_key():
        st.info("Enter your GOOGLE_API_KEY to enable AI responses.")
        st.session_state.GOOGLE_API_KEY_UI = st.text_input("GOOGLE_API_KEY", type="password")
    st.session_state.enable_tts = st.checkbox("Enable Patient Voice", value=st.session_state.enable_tts)
    st.markdown("---")

def page_registration():
    st.markdown("<h1 style='text-align:center;color:#111827;'>Psychiatric Interview Simulation</h1>", unsafe_allow_html=True)
    tab_login, tab_register = st.tabs(["Login", "Register"])

    with tab_login:
        u = st.text_input("Username")
        if st.button("Login", type="primary"):
            data = load_user_profile(u)
            if data:
                st.session_state.user = data
                st.session_state.page = "menu"
                st.session_state.session_id = None
                st.success(f"Welcome back, {data['nickname']}!")
                st.rerun()
            else:
                st.error("User not found. Please register first.")

    with tab_register:
        c1, c2 = st.columns(2)
        with c1:
            first = st.text_input("First Name*")
            last = st.text_input("Last Name*")
            nick = st.text_input("Nickname*")
        with c2:
            email = st.text_input("Email (optional)")
            photo = st.file_uploader("Upload Your Photo*", type=["jpg","jpeg","png","webp"])

        if st.button("Register", type="primary"):
            if first and last and nick and photo:
                username = nick.lower().replace(" ", "_")
                ext = os.path.splitext(photo.name)[1].lower() or ".jpg"
                photo_path = os.path.join(USER_DIR, f"{username}_photo{ext}")
                with open(photo_path, "wb") as f:
                    f.write(photo.getbuffer())
                user_data = {
                    "username": username,
                    "first_name": first,
                    "last_name": last,
                    "nickname": nick,
                    "email": email,
                    "photo_path": photo_path,
                    "registration_date": datetime.datetime.now().isoformat()
                }
                save_user_profile(username, user_data)
                save_user_to_excel(user_data)
                st.session_state.user = user_data
                st.session_state.page = "menu"
                st.success("Registration successful.")
                st.rerun()
            else:
                st.error("Please fill all required fields (*) and upload a photo.")

def page_menu():
    user = st.session_state.user
    st.markdown(
        f"""
        <div style='background:rgba(255,255,255,0.92);padding:16px;border-radius:12px;border:1px solid #eee'>
            <h2 style='margin:0'>Welcome, {user['nickname']}!</h2>
            <p style='margin:4px 0 0 0'>Select a patient and interview stage to begin.</p>
        </div>
        """,
        unsafe_allow_html=True
    )

    c1, c2 = st.columns(2)
    with c1:
        st.subheader("Aliye Seker — MDD")
        if os.path.exists(MDD_PERSONA["photo"]):
            st.image(MDD_PERSONA["photo"], use_container_width=True)
        else:
            st.info(f"{MDD_PERSONA['photo']} not found")
        if st.button("Select Aliye Seker"):
            st.session_state.selected_persona = MDD_PERSONA
            st.rerun()
    with c2:
        st.subheader("Feride Deniz — Schizophrenia")
        if os.path.exists(SCZ_PERSONA["photo"]):
            st.image(SCZ_PERSONA["photo"], use_container_width=True)
        else:
            st.info(f"{SCZ_PERSONA['photo']} not found")
        if st.button("Select Feride Deniz"):
            st.session_state.selected_persona = SCZ_PERSONA
            st.rerun()

    if st.session_state.selected_persona:
        st.markdown("---")
        st.info(f"Selected patient: **{st.session_state.selected_persona['name']}**")
        colp1, colp2 = st.columns(2)
        with colp1:
            if st.button("Part 1"):
                st.session_state.selected_part = "Part 1"
                st.session_state.session_id = str(uuid.uuid4())[:8]
                st.session_state.conversation = []
                st.session_state.awaiting_permission = False
                st.session_state.page = "interview"
                st.rerun()
        with colp2:
            if st.button("Part 2"):
                st.session_state.selected_part = "Part 2"
                st.session_state.session_id = str(uuid.uuid4())[:8]
                st.session_state.conversation = []
                st.session_state.awaiting_permission = False
                st.session_state.page = "interview"
                st.rerun()

    st.markdown("---")
    if st.button("Logout"):
        st.session_state.clear()
        ensure_state_defaults()
        st.rerun()

def page_interview():
    persona = st.session_state.selected_persona
    part = st.session_state.selected_part
    sid = st.session_state.session_id

    with st.sidebar:
        sidebar_patient_info(persona)

    st.markdown(
        f"""
        <div style='background:linear-gradient(135deg,#667eea,#764ba2);padding:14px;border-radius:12px;color:#fff'>
            <h3 style='margin:0'>Interview • {persona['name']} — {part}</h3>
            <small>Session ID: {sid}</small>
        </div>
        """,
        unsafe_allow_html=True
    )

    # Conversation (render last MAX_MSG)
    st.markdown("### Conversation History")
    patient_photo = persona.get("photo", "")
    patient_avatar_b64 = load_avatar_b64(patient_photo)
    student_photo = st.session_state.user.get("photo_path", "")
    student_avatar_b64 = load_avatar_b64(student_photo)

    history = st.session_state.conversation[-MAX_MSG:]
    for role, msg, ts, source in history:
        if role == "Student":
            col1, col2 = st.columns([1, 9])
            with col1:
                if student_avatar_b64:
                    st.markdown(
                        f'<img src="data:image/jpeg;base64,{student_avatar_b64}" style="width:100%;border-radius:50%;border:3px solid #3b82f6;">',
                        unsafe_allow_html=True
                    )
                else:
                    st.markdown("👤")
            with col2:
                st.markdown(
                    f"""<div style='background:#e0f2fe;padding:12px;border-radius:12px;border-left:4px solid #0284c7;'>
                    <div style='font-weight:600;color:#0c4a6e;margin-bottom:4px;'>You</div>
                    <div style='color:#075985;'>{msg}</div>
                    <div style='color:#64748b;font-size:11px;margin-top:4px;'>{ts}</div>
                    </div>""",
                    unsafe_allow_html=True
                )
        else:
            col1, col2 = st.columns([1, 9])
            with col1:
                if patient_avatar_b64:
                    st.markdown(
                        f'<img src="data:image/jpeg;base64,{patient_avatar_b64}" style="width:100%;border-radius:50%;border:3px solid #8b5cf6;">',
                        unsafe_allow_html=True
                    )
                else:
                    st.markdown("👤")
            with col2:
                badge = "STANDARDIZED" if source == "db" else "AI"
                badge_color = "#10b981" if source == "db" else "#8b5cf6"
                st.markdown(
                    f"""<div style='background:#f3e8ff;padding:12px;border-radius:12px;border-left:4px solid #a855f7;'>
                    <div style='font-weight:600;color:#6b21a8;margin-bottom:4px;'>
                        {persona['name']} 
                        <span style='font-size:10px;background:{badge_color};color:white;padding:2px 8px;border-radius:10px;margin-left:8px;font-weight:500;'>{badge}</span>
                    </div>
                    <div style='color:#7c3aed;'>{msg}</div>
                    <div style='color:#64748b;font-size:11px;margin-top:4px;'>{ts}</div>
                    </div>""",
                    unsafe_allow_html=True
                )
        st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)

    st.markdown("---")
    st.subheader("Your Input")

    input_mode = st.radio("Choose input method", ["Text", "Voice (Automatic)"], horizontal=True, key="input_mode_radio")

    if input_mode == "Text":
        col1, col2 = st.columns([5,1])
        with col1:
            user_text = st.text_input("Type your question or statement:", key="text_in")
        with col2:
            send = st.button("Send", type="primary")
        if send and (user_text or "").strip():
            handle_turn(user_text.strip())
            st.rerun()  # rerun OK for Text
    else:
        st.info("Press & hold the mic to record. Release to stop. The transcript will be sent automatically.")
        # Persistent mic (no rerun)
        transcript = persistent_recorder(language="en-US")  # change to "tr-TR" if needed
        if transcript and transcript != st.session_state.get("pending_voice_input", ""):
            st.session_state.pending_voice_input = transcript
            handle_turn(transcript)   # no st.rerun()

    st.markdown("---")
    c1, c2 = st.columns(2)
    with c1:
        if st.button("End Session"):
            st.session_state.page = "evaluation"
            st.rerun()
    with c2:
        if st.button("Back to Menu"):
            st.session_state.page = "menu"
            st.rerun()

def handle_turn(user_input: str):
    persona = st.session_state.selected_persona
    part = st.session_state.selected_part

    log_line("Student", user_input, "")

    photo = persona.get("photo", "")
    if os.path.exists(photo):
        show_avatar(photo, speaking=False, placeholder=st.session_state.avatar_placeholder)

    # shorter delay for UX feel
    # time.sleep(0.05)

    ans_db = qa_lookup(user_input, persona["name"], part)
    if ans_db:
        ans, source = ans_db, "db"
    else:
        ans, source = llm_reply(persona, part, user_input), "ai"

    if os.path.exists(photo):
        show_avatar(photo, speaking=True, placeholder=st.session_state.avatar_placeholder)

    if st.session_state.enable_tts:
        speak_browser(ans)
        # time.sleep(0.2)

    if os.path.exists(photo):
        show_avatar(photo, speaking=False, placeholder=st.session_state.avatar_placeholder)

    log_line("Patient", ans, source)

def page_evaluation():
    db = sum(1 for r in st.session_state.conversation if r[0]=="Patient" and r[3]=="db")
    ai = sum(1 for r in st.session_state.conversation if r[0]=="Patient" and r[3]=="ai")
    st.success("Interview Completed. You can start a new session from the menu.")
    c1, c2 = st.columns(2)
    with c1:
        st.metric("Standardized Responses", db)
    with c2:
        st.metric("AI Responses", ai)
    if st.button("Back to Menu", type="primary"):
        st.session_state.page = "menu"
        st.rerun()

# =============================
# MAIN
# =============================
def main():
    if st.session_state.page == "registration":
        page_registration()
        return

    if not st.session_state.user:
        st.session_state.page = "registration"
        st.rerun()

    if st.session_state.page == "menu":
        page_menu()
    elif st.session_state.page == "interview":
        if not (st.session_state.selected_persona and st.session_state.selected_part and st.session_state.session_id):
            st.session_state.page = "menu"
            st.rerun()
        page_interview()
    elif st.session_state.page == "evaluation":
        page_evaluation()

    st.markdown(
        """
        <div style="margin-top:24px;padding:12px;border-top:1px solid #e5e7eb;color:#6b7280">
            Developed by Dr. Volkan OBAN — 2025
        </div>
        """,
        unsafe_allow_html=True
    )

if __name__ == "__main__":
    main()
